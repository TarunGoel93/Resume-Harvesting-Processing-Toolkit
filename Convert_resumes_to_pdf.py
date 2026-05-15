"""
Resume Link -> PDF URL Converter
Converts each resume link to PDF via pdf.co and writes the result URL
back into the Excel file immediately after each file is processed.

A new column "PDF Link" is added to the Excel file.
The file is saved after every row so progress is never lost.

Requirements:
    pip install openpyxl requests

Usage:
    1. Set API_KEY below
    2. Place Resumes-2K.xlsx in the same folder
    3. python convert_resumes_to_pdf.py
"""

import time
import requests
import openpyxl
from openpyxl.styles import Font
from pathlib import Path

# --- CONFIGURATION -----------------------------------------------------------

API_KEY    = "tushargoel921@gmail.com_ENKpJrVDbGv9HobWrbhG0XODWjPgIQkyVqUKS7fp0XElLl8eom704hHAss7F7zUs"   # <- paste your pdf.co API key here
EXCEL_FILE = "Resumes-2K.xlsx"
DELAY_SEC  = 0.5
MAX_RETRIES = 3

ENDPOINT_DOC = "https://api.pdf.co/v1/pdf/convert/from/doc"
ENDPOINT_URL = "https://api.pdf.co/v1/pdf/convert/from/url"

# -----------------------------------------------------------------------------


def api_headers():
    return {"x-api-key": API_KEY, "Content-Type": "application/json"}


def drive_file_id(url):
    if "/file/d/" in url:
        return url.split("/file/d/")[1].split("/")[0]
    return None


def docs_export_url(url):
    if "docs.google.com/document/d/" in url:
        doc_id = url.split("/document/d/")[1].split("/")[0]
        return f"https://docs.google.com/document/d/{doc_id}/export?format=pdf"
    return None


def call_pdfco(endpoint, source_url, output_name):
    """Call pdf.co and return the result PDF URL, or None on failure."""
    payload = {"url": source_url, "name": output_name, "async": False}

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = requests.post(endpoint, headers=api_headers(), json=payload, timeout=90)
            data = resp.json()

            if resp.status_code == 200 and not data.get("error"):
                return data.get("url")

            msg = data.get("message", resp.text)
            print(f"\n    [!] API error (attempt {attempt}/{MAX_RETRIES}): {msg}", end="")
            if attempt < MAX_RETRIES:
                time.sleep(2 ** attempt)

        except requests.RequestException as e:
            print(f"\n    [!] Network error (attempt {attempt}/{MAX_RETRIES}): {e}", end="")
            if attempt < MAX_RETRIES:
                time.sleep(2 ** attempt)

    return None


def get_pdf_url(file_name, file_type, raw_url):
    """
    Route each file to the right strategy and return a PDF URL.

    PDF on Drive       -> upload to pdf.co storage, get back a hosted PDF URL
    DOC/DOCX on Docs   -> Google export URL -> pdf.co converts it
    DOC/DOCX on Drive  -> direct Drive URL  -> pdf.co /from/doc converts it
    """
    ft = (file_type or "").upper()
    stem = Path(file_name).stem
    output_name = f"{stem}.pdf"

    if ft == "PDF":
        # Upload the Drive file to pdf.co storage so it gets a permanent URL
        fid = drive_file_id(raw_url)
        if not fid:
            return None
        direct_url = f"https://drive.google.com/uc?export=download&id={fid}"
        # Use /url endpoint with the direct download link — pdf.co will host it
        return call_pdfco(ENDPOINT_URL, direct_url, output_name)

    if "docs.google.com/document" in raw_url:
        export_url = docs_export_url(raw_url)
        if not export_url:
            return None
        return call_pdfco(ENDPOINT_URL, export_url, output_name)

    # Raw DOC/DOCX on Drive
    fid = drive_file_id(raw_url)
    direct_url = f"https://drive.google.com/uc?export=download&id={fid}" if fid else raw_url
    return call_pdfco(ENDPOINT_DOC, direct_url, output_name)


def setup_pdf_column(ws):
    """
    Find or create the 'PDF Link' column. Returns the column index (1-based).
    Also makes the header bold.
    """
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    if "PDF Link" in headers:
        col = headers.index("PDF Link") + 1
    else:
        col = ws.max_column + 1
        cell = ws.cell(1, col)
        cell.value = "PDF Link"
        cell.font = Font(bold=True)

    return col


def main():
    if API_KEY == "YOUR_PDF_CO_API_KEY":
        raise SystemExit("Please set your pdf.co API key in the API_KEY variable.")

    excel_path = Path(EXCEL_FILE)
    if not excel_path.exists():
        raise SystemExit(f"Excel file not found: {excel_path.resolve()}")

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    pdf_col = setup_pdf_column(ws)
    total   = ws.max_row - 1  # exclude header

    print(f"Processing {total} rows. PDF links will be saved to column '{ws.cell(1, pdf_col).value}'.")
    print(f"Excel file: {excel_path.resolve()}\n")

    success = skip = fail = 0

    for row in range(2, ws.max_row + 1):
        idx       = row - 1
        file_name = ws.cell(row, 1).value
        file_type = ws.cell(row, 2).value
        link_cell = ws.cell(row, 5)
        pdf_cell  = ws.cell(row, pdf_col)

        # Skip if already processed
        if pdf_cell.value and pdf_cell.value.startswith("http"):
            print(f"[{idx}/{total}] SKIP  {file_name}")
            skip += 1
            continue

        if not (link_cell.hyperlink and link_cell.hyperlink.target):
            print(f"[{idx}/{total}] SKIP  {file_name}  (no hyperlink)")
            skip += 1
            continue

        raw_url = link_cell.hyperlink.target
        print(f"[{idx}/{total}] {file_name} ({file_type})", end="  ->  ", flush=True)

        pdf_url = get_pdf_url(str(file_name or f"row_{row}"), file_type, raw_url)

        if pdf_url:
            # Store as a clickable hyperlink in the cell
            pdf_cell.value = pdf_url
            pdf_cell.hyperlink = pdf_url
            pdf_cell.font = Font(color="0563C1", underline="single")
            print("OK")
            success += 1
        else:
            pdf_cell.value = "FAILED"
            print("FAIL")
            fail += 1

        # Save after every row so progress is never lost
        wb.save(excel_path)

        time.sleep(DELAY_SEC)

    print(f"\n{'─' * 50}")
    print(f"Converted : {success}")
    print(f"Skipped   : {skip}")
    print(f"Failed    : {fail}")
    print(f"Saved to  : {excel_path.resolve()}")


if __name__ == "__main__":
    main()