"""
Scan a Google Drive folder and generate an Excel (.xlsx) file
with clickable hyperlinks to every PDF / DOC / DOCX file found.

Setup:
1. pip install google-api-python-client google-auth-httplib2 google-auth-oauthlib openpyxl
2. Go to console.cloud.google.com → Enable Google Drive API
3. Create OAuth 2.0 Desktop credentials → download as credentials.json
4. Place credentials.json next to this script
5. Run: python drive_to_excel.py
"""

import sys
from datetime import datetime
from pathlib import Path

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CONFIG ────────────────────────────────────────────────────────────────────
DRIVE_FOLDER_NAME = "K"   # ← exact name of your Drive folder
OUTPUT_XLSX       = "drive_file_index.xlsx"
# ─────────────────────────────────────────────────────────────────────────────

SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]

SUPPORTED_MIMES = {
    "application/pdf":                                                          "PDF",
    "application/msword":                                                       "DOC",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document":  "DOCX",
}

# ── Column definitions: (header, width) ──────────────────────────────────────
COLUMNS = [
    ("#",            6),
    ("File Name",   55),
    ("Type",        10),
    ("Size (KB)",   12),
    ("Modified",    20),
    ("Open File",   20),
]

# ── Styling ───────────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY_FONT    = Font(name="Arial", size=10)
LINK_FONT    = Font(name="Arial", size=10, color="0563C1", underline="single")
ALT_FILL     = PatternFill("solid", fgColor="EBF3FB")

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")


# ── Auth ──────────────────────────────────────────────────────────────────────
def authenticate():
    creds = None
    token_path = Path("token.json")
    creds_path = Path("credentials.json")

    if not creds_path.exists():
        print("ERROR: credentials.json not found.")
        print("Download it from Google Cloud Console → APIs & Services → Credentials.")
        sys.exit(1)

    if token_path.exists():
        creds = Credentials.from_authorized_user_file(token_path, SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(creds_path, SCOPES)
            creds = flow.run_local_server(port=0)
        token_path.write_text(creds.to_json())

    return build("drive", "v3", credentials=creds)


# ── Drive helpers ─────────────────────────────────────────────────────────────
def find_folder_id(service, folder_name):
    q = (
        f"name='{folder_name}' "
        "and mimeType='application/vnd.google-apps.folder' "
        "and trashed=false"
    )
    res = service.files().list(q=q, fields="files(id,name)").execute()
    files = res.get("files", [])
    if not files:
        print(f"ERROR: No Drive folder named '{folder_name}' found.")
        print("Check the DRIVE_FOLDER_NAME setting at the top of the script.")
        sys.exit(1)
    return files[0]["id"]


def list_files(service, folder_id):
    mime_filter = " or ".join(
        f"mimeType='{m}'" for m in SUPPORTED_MIMES
    )
    q = f"'{folder_id}' in parents and ({mime_filter}) and trashed=false"
    fields = "nextPageToken, files(id, name, mimeType, size, modifiedTime, webViewLink)"

    results, page_token = [], None
    while True:
        resp = service.files().list(
            q=q, fields=fields, pageSize=1000, pageToken=page_token
        ).execute()
        results.extend(resp.get("files", []))
        page_token = resp.get("nextPageToken")
        if not page_token:
            break

    return sorted(results, key=lambda f: f["name"].lower())


# ── Excel builder ─────────────────────────────────────────────────────────────
def style_header(ws):
    for col_idx, (header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 22


def style_cell(cell, alt_row, link=False, center=False):
    cell.font      = LINK_FONT if link else BODY_FONT
    cell.border    = BORDER
    cell.alignment = CENTER if center else LEFT
    if alt_row:
        cell.fill = ALT_FILL


def build_excel(files):
    wb = Workbook()
    ws = wb.active
    ws.title = "Drive File Index"
    ws.freeze_panes = "A2"

    style_header(ws)

    for i, f in enumerate(files, start=1):
        row     = i + 1
        alt     = i % 2 == 0
        name    = f["name"]
        ftype   = SUPPORTED_MIMES.get(f["mimeType"], "?")
        size_kb = round(int(f.get("size", 0)) / 1024, 1)
        modified = datetime.fromisoformat(
            f["modifiedTime"].replace("Z", "+00:00")
        ).strftime("%d %b %Y  %H:%M")
        link = f.get("webViewLink", "")

        values = [i, name, ftype, size_kb, modified]
        for col_idx, val in enumerate(values, start=1):
            c = ws.cell(row=row, column=col_idx, value=val)
            style_cell(c, alt, center=(col_idx in (1, 3, 4)))

        # Clickable hyperlink in column 6
        link_cell = ws.cell(row=row, column=6, value="📂 Open")
        link_cell.hyperlink = link
        style_cell(link_cell, alt, link=True, center=True)

        ws.row_dimensions[row].height = 18

    # Summary row
    summary_row = len(files) + 2
    ws.cell(row=summary_row, column=1, value=f"Total: {len(files)} file(s)").font = Font(
        name="Arial", bold=True, italic=True, size=10, color="555555"
    )

    wb.save(OUTPUT_XLSX)
    return OUTPUT_XLSX


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print("Authenticating with Google Drive ...")
    service = authenticate()

    print(f"Looking for folder: '{DRIVE_FOLDER_NAME}' ...")
    folder_id = find_folder_id(service, DRIVE_FOLDER_NAME)

    print("Fetching file list ...")
    files = list_files(service, folder_id)

    if not files:
        print("No PDF / DOC / DOCX files found in that folder.")
        sys.exit(0)

    print(f"Found {len(files)} file(s). Building Excel ...")
    out = build_excel(files)

    print(f"\n✅ Done!  →  {out}")
    print("Open the Excel file — click '📂 Open' in any row to open that file in Drive.\n")


if __name__ == "__main__":
    main()

